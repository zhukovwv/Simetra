from openpyxl.reader.excel import load_workbook
from pydantic import FilePath

from database import get_async_session
from models.models import GPSData

import datetime


async def load_data_from_excel(file_path: FilePath) -> None:
    async for session in get_async_session():
        # Подключение к Excel файлу
        wb = load_workbook(filename=file_path)
        ws = wb.active
        try:
            # Проход по строкам и столбцам в Excel и добавление данных в базу данных
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Replace commas with periods in latitude and longitude values
                longitude = row[1]
                latitude = row[2]

                # Преобразование временной метки строки в дату и время без информации о часовом поясе
                gps_time = datetime.datetime.strptime(row[4], '%Y-%m-%dT%H:%M:%S%z').replace(tzinfo=None)

                gps_data = GPSData(
                    id=row[0],
                    location=f"POINT({longitude} {latitude})",
                    speed=row[3],
                    gps_time=gps_time,
                    vehicle_id=row[5]
                )
                session.add(gps_data)

            # Сохранение изменений в базе данных
            await session.commit()
            print("Данные успешно загружены в базу данных.")
        except Exception as e:
            # В случае возникновения ошибки откатываем транзакцию
            await session.rollback()
            print(f"Произошла ошибка: {str(e)}")
        finally:
            await session.close()
